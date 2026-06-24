import json
import logging
import os
import sqlite3
import sys
from pathlib import Path
from tkinter import Tk, filedialog

import ezsheets
import openpyxl
import pymsgbox

from Function import Functions


class SQLiteBillingDatabase:
    """SQLite storage that keeps the account dictionary API used by the TUI."""

    def __init__(self, database_path, legacy_database_paths=None, writeback=False):
        self.database_path = Path(database_path)
        self.legacy_database_paths = [Path(path) for path in (legacy_database_paths or [])]
        self.writeback = writeback
        self.connection = None
        self._accounts_cache = None

    def __enter__(self):
        self.database_path.parent.mkdir(parents=True, exist_ok=True)
        self.connection = sqlite3.connect(str(self.database_path))
        self.connection.row_factory = sqlite3.Row
        self.connection.execute("PRAGMA foreign_keys = ON")
        self._create_schema()
        self._migrate_legacy_data()
        return self

    def __exit__(self, exc_type, exc_value, traceback):
        try:
            if exc_type is None and self.writeback and self._accounts_cache is not None:
                self._save_accounts(self._accounts_cache)
        finally:
            if self.connection is not None:
                self.connection.close()

    def __contains__(self, key):
        return key == "accounts"

    def __getitem__(self, key):
        if key == "accounts":
            return self._load_accounts()
        raise KeyError(key)

    def __setitem__(self, key, value):
        if key != "accounts":
            raise KeyError(key)
        self._save_accounts(value)

    def get(self, key, default=None):
        if key == "accounts":
            return self._load_accounts()
        if key == "contas":
            return default
        return default

    def _create_schema(self):
        with self.connection:
            self.connection.execute(
                """
                CREATE TABLE IF NOT EXISTS accounts (
                    name TEXT PRIMARY KEY,
                    code TEXT NOT NULL,
                    created_at TEXT NOT NULL,
                    billing_items TEXT NOT NULL DEFAULT '{}'
                )
                """
            )
            self.connection.execute(
                """
                CREATE TABLE IF NOT EXISTS purchases (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    nc TEXT NOT NULL UNIQUE,
                    account_name TEXT NOT NULL,
                    item TEXT NOT NULL,
                    price TEXT,
                    date TEXT,
                    FOREIGN KEY (account_name) REFERENCES accounts(name) ON DELETE CASCADE
                )
                """
            )
            self.connection.execute(
                """
                CREATE INDEX IF NOT EXISTS idx_purchases_account_name
                ON purchases(account_name)
                """
            )

    def _load_accounts(self):
        if self._accounts_cache is not None:
            return self._accounts_cache

        accounts = {}
        account_rows = self.connection.execute(
            """
            SELECT name, code, created_at, billing_items
            FROM accounts
            ORDER BY created_at DESC, name
            """
        ).fetchall()

        for account_row in account_rows:
            try:
                billing_items = json.loads(account_row["billing_items"] or "{}")
            except json.JSONDecodeError:
                logging.warning("Invalid billing_items JSON for account %s.", account_row["name"])
                billing_items = {}

            purchase_rows = self.connection.execute(
                """
                SELECT nc, item, price, date
                FROM purchases
                WHERE account_name = ?
                ORDER BY id
                """,
                (account_row["name"],),
            ).fetchall()

            accounts[account_row["name"]] = {
                "code": account_row["code"],
                "created_at": account_row["created_at"],
                "billing_items": billing_items,
                "purchase_history": [
                    {
                        "nc": purchase_row["nc"],
                        "item": purchase_row["item"],
                        "price": purchase_row["price"],
                        "date": purchase_row["date"],
                    }
                    for purchase_row in purchase_rows
                ],
            }

        self._accounts_cache = accounts
        return accounts

    def _save_accounts(self, accounts):
        with self.connection:
            self.connection.execute("DELETE FROM purchases")
            self.connection.execute("DELETE FROM accounts")

            for account_name, details in accounts.items():
                billing_items = details.get("billing_items", details.get("itens_faturamento", {})) or {}
                self.connection.execute(
                    """
                    INSERT INTO accounts (name, code, created_at, billing_items)
                    VALUES (?, ?, ?, ?)
                    """,
                    (
                        account_name,
                        details.get("code", details.get("Codigo", "")),
                        details.get("created_at", details.get("criada_em", "")),
                        json.dumps(billing_items, ensure_ascii=False),
                    ),
                )

                history = details.get("purchase_history", details.get("historico_compras", [])) or []
                for purchase in history:
                    nc_code = purchase.get("nc")
                    if not nc_code:
                        continue

                    self.connection.execute(
                        """
                        INSERT OR REPLACE INTO purchases (nc, account_name, item, price, date)
                        VALUES (?, ?, ?, ?, ?)
                        """,
                        (
                            nc_code,
                            account_name,
                            purchase.get("item", ""),
                            purchase.get("price", purchase.get("preco")),
                            purchase.get("date", purchase.get("data")),
                        ),
                    )

        self._accounts_cache = accounts

    def _has_accounts(self):
        row = self.connection.execute("SELECT 1 FROM accounts LIMIT 1").fetchone()
        return row is not None

    def _migrate_legacy_data(self):
        if self._has_accounts():
            return

        legacy_accounts = self._read_legacy_shelve_accounts()
        if legacy_accounts:
            self._save_accounts(legacy_accounts)
            logging.info("Migrated %s billing accounts from shelve to SQLite.", len(legacy_accounts))

    def _read_legacy_shelve_accounts(self):
        for legacy_path in self.legacy_database_paths:
            if not self._legacy_shelve_exists(legacy_path):
                continue

            try:
                import shelve

                with shelve.open(str(legacy_path)) as database:
                    accounts = database.get("accounts", database.get("contas", {}))
                    if accounts:
                        return dict(accounts)
            except Exception as exc:
                logging.warning(
                    "Could not migrate legacy shelve database %s: %s",
                    legacy_path,
                    exc,
                    exc_info=True,
                )

        return {}

    def _legacy_shelve_exists(self, legacy_path):
        suffixes = ["", ".bak", ".dat", ".dir", ".db"]
        return any(legacy_path.with_name(legacy_path.name + suffix).exists() for suffix in suffixes)


class Billing:
    def __init__(self, data_folder=None):
        self.data_folder = data_folder or (Path.home() / "SISTEMA_LF" / "Data")
        self.helpers = Functions()
        self.database_path = self.data_folder / "billing_accounts.sqlite3"
        self.legacy_database_paths = [
            self.data_folder / "billing_accounts",
            self.data_folder / "faturamento_contas",
        ]

    def _open_database(self, writeback=False):
        return SQLiteBillingDatabase(
            self.database_path,
            self.legacy_database_paths,
            writeback=writeback,
        )

    def _get_accounts(self, database):
        return database.get("accounts", database.get("contas", {}))

    def _ensure_accounts(self, database):
        if "accounts" not in database:
            database["accounts"] = database.get("contas", {})
        return database["accounts"]

    def monitor_customers(self):
        logging.info("Customer monitoring started.")
        try:
            with self._open_database() as database:
                all_accounts = self._get_accounts(database)
                if not all_accounts:
                    self._empty_screen("Customers")
                    return

                ordered_accounts = sorted(
                    list(all_accounts.items()),
                    key=lambda item: item[1].get("created_at", item[1].get("criada_em", "")),
                    reverse=True,
                )
                pages = [ordered_accounts[i : i + 10] for i in range(0, len(ordered_accounts), 10)]
                current_page = 0

                while True:
                    os.system("cls" if os.name == "nt" else "clear")
                    print("--- CUSTOMER MONITOR ---")
                    print(f"Page: {current_page + 1}/{max(1, len(pages))}\n" + "-" * 60)

                    for account_name, details in pages[current_page]:
                        created = details.get("created_at", details.get("criada_em", "N/A"))
                        billing_items = details.get("billing_items", details.get("itens_faturamento", {}))
                        customer = billing_items.get("customer_name", billing_items.get("nome_cliente", "N/A"))
                        code = details.get("code", details.get("Codigo", "N/A"))
                        print(
                            f"[{code}] Account: {account_name:<12} | "
                            f"Customer: {customer:<18} | Created: {created}"
                        )

                    print("-" * 60)
                    print(" [1] Next Page | [2] Previous Page | [S] Back")
                    command = input("Command: ").strip().lower()
                    if command == "1" and current_page < len(pages) - 1:
                        current_page += 1
                    elif command == "2" and current_page > 0:
                        current_page -= 1
                    elif command == "s":
                        break
        except Exception as exc:
            logging.error("Customer monitor error: %s", exc, exc_info=True)

    def monitor_purchases(self):
        logging.info("Global purchase monitoring started.")
        try:
            with self._open_database() as database:
                all_accounts = self._get_accounts(database)
                if not all_accounts:
                    self._empty_screen("Purchases")
                    return

                global_purchases = []
                for account_name, details in all_accounts.items():
                    history = details.get("purchase_history", details.get("historico_compras", []))
                    for purchase in history:
                        global_purchases.append(
                            {
                                "nc": purchase.get("nc", "N/NC"),
                                "account": account_name,
                                "item": purchase["item"],
                                "price": purchase.get("price", purchase.get("preco")),
                                "date": purchase.get("date", purchase.get("data")),
                            }
                        )

                if not global_purchases:
                    self._empty_screen("Purchases")
                    return

                global_purchases = sorted(global_purchases, key=lambda item: item["date"], reverse=True)
                pages = [global_purchases[i : i + 10] for i in range(0, len(global_purchases), 10)]
                current_page = 0

                while True:
                    os.system("cls" if os.name == "nt" else "clear")
                    print("--- GLOBAL PURCHASE HISTORY ---")
                    print(f"Page: {current_page + 1}/{max(1, len(pages))}\n" + "-" * 70)

                    for purchase in pages[current_page]:
                        print(
                            f"[{purchase['nc']:<5}] Item: {purchase['item']:<15} | "
                            f"$ {purchase['price']:<8} | Account: {purchase['account']:<10} | "
                            f"Date: {purchase['date']}"
                        )

                    print("-" * 70)
                    print(" [1] Next Page | [2] Previous Page | [S] Back")
                    command = input("Command: ").strip().lower()
                    if command == "1" and current_page < len(pages) - 1:
                        current_page += 1
                    elif command == "2" and current_page > 0:
                        current_page -= 1
                    elif command == "s":
                        break
        except Exception as exc:
            logging.error("Purchase monitor error: %s", exc, exc_info=True)

    def _empty_screen(self, item_type):
        os.system("cls" if os.name == "nt" else "clear")
        print(f"--- Empty {item_type} History ---")
        input("\nNo records found. Press Enter...")

    def create_billing_account(self):
        try:
            with self._open_database(writeback=True) as database:
                all_accounts = self._ensure_accounts(database)

                account_name = pymsgbox.prompt("New account name (example: Cash_A):", "Create Account")
                if account_name and account_name.strip():
                    account_name = account_name.strip()
                    if account_name in all_accounts:
                        pymsgbox.alert("This account already exists.", "Error")
                        return

                    highest_id = -1
                    for account in all_accounts.values():
                        code = account.get("code", account.get("Codigo", "LF0"))
                        number = int(code.replace("LF", ""))
                        highest_id = max(highest_id, number)

                    new_code = f"LF{highest_id + 1}"
                    all_accounts[account_name] = {
                        "code": new_code,
                        "created_at": self.helpers.generate_timestamp(),
                        "billing_items": {
                            "age": "Not provided",
                            "customer_name": "Not provided",
                            "customer_address": "Not provided",
                        },
                        "purchase_history": [],
                    }
                    database["accounts"] = all_accounts
                    pymsgbox.alert(
                        f"Account '{account_name}' created with code {new_code}.",
                        "Success",
                    )
        except Exception as exc:
            logging.error("Create account error: %s", exc, exc_info=True)

    def show_account_details(self):
        try:
            with self._open_database() as database:
                all_accounts = self._get_accounts(database)
                selected_account = pymsgbox.prompt("Enter the account name:")
                if selected_account in all_accounts:
                    details = all_accounts[selected_account]
                    billing_items = details.get("billing_items", details.get("itens_faturamento", {}))
                    items = "\n".join([f"{key}: {value}" for key, value in billing_items.items()])
                    info = (
                        f"Account: {selected_account}\n"
                        f"Code: {details.get('code', details.get('Codigo'))}\n"
                        f"Created: {details.get('created_at', details.get('criada_em'))}\n\n"
                        f"Fields:\n{items}"
                    )
                    pymsgbox.alert(info, "Details")
                else:
                    pymsgbox.alert("Account not found.", "Error")
        except Exception as exc:
            logging.error("Account details error: %s", exc, exc_info=True)

    def add_purchase(self, account_name, item, price):
        try:
            with self._open_database(writeback=True) as database:
                all_accounts = self._ensure_accounts(database)

                if account_name in all_accounts:
                    highest_nc = -1
                    for account in all_accounts.values():
                        history = account.get("purchase_history", account.get("historico_compras", []))
                        for purchase in history:
                            number = int(purchase["nc"].replace("NC", ""))
                            highest_nc = max(highest_nc, number)

                    new_nc = f"NC{highest_nc + 1}"
                    all_accounts[account_name].setdefault("purchase_history", []).append(
                        {
                            "nc": new_nc,
                            "item": item,
                            "price": price,
                            "date": self.helpers.generate_timestamp(),
                        }
                    )
                    database["accounts"] = all_accounts
                    pymsgbox.alert(f"Purchase posted. Code: {new_nc}", "Success")
                else:
                    pymsgbox.alert("Account is not registered.", "Error")
        except Exception as exc:
            logging.error("Add purchase error: %s", exc, exc_info=True)

    def delete_purchase(self):
        try:
            with self._open_database(writeback=True) as database:
                all_accounts = self._ensure_accounts(database)
                target = pymsgbox.prompt("Enter the NC code for the item to delete:")
                if not target:
                    return
                target = target.strip().upper()

                found = False
                for details in all_accounts.values():
                    history = details.get("purchase_history", details.get("historico_compras", []))
                    for purchase in list(history):
                        if purchase.get("nc") == target:
                            confirmation = pymsgbox.prompt(
                                f"Type DELETE to confirm deletion of {purchase['item']}:"
                            )
                            if confirmation == "DELETE":
                                history.remove(purchase)
                                details["purchase_history"] = history
                                pymsgbox.alert("Item removed.", "Success")
                                found = True
                            break

                database["accounts"] = all_accounts
                if not found:
                    pymsgbox.alert("Code not found.", "Error")
        except Exception as exc:
            logging.error("Delete purchase error: %s", exc, exc_info=True)

    def edit_account(self):
        try:
            with self._open_database(writeback=True) as database:
                all_accounts = self._ensure_accounts(database)
                selected_account = pymsgbox.prompt("Account to update:")
                if selected_account in all_accounts:
                    items = all_accounts[selected_account].setdefault("billing_items", {})
                    if not items and "itens_faturamento" in all_accounts[selected_account]:
                        items.update(all_accounts[selected_account]["itens_faturamento"])
                    keys = list(items.keys())

                    while True:
                        os.system("cls" if os.name == "nt" else "clear")
                        print(f"--- EDIT ACCOUNT: {selected_account} ---")
                        for index, key in enumerate(keys):
                            print(f" [{index}] {key.upper()}: {items[key]}")
                        print(" [S] Exit")

                        option = input("\nChoose the field index to update: ").strip()
                        if option.lower() == "s":
                            break
                        if option.isdigit() and int(option) < len(keys):
                            target_key = keys[int(option)]
                            new_value = pymsgbox.prompt(
                                f"New value for {target_key}:",
                                default=items[target_key],
                            )
                            if new_value is not None:
                                items[target_key] = new_value.strip()

                    database["accounts"] = all_accounts
                else:
                    pymsgbox.alert("Account does not exist.", "Error")
        except Exception as exc:
            logging.error("Edit account error: %s", exc, exc_info=True)

class SpreadsheetExport:
    def __init__(self, spreadsheets_folder, data_folder=None):
        self.spreadsheets_folder = Path(spreadsheets_folder)
        self.data_folder = Path(data_folder) if data_folder else (Path.home() / "SISTEMA_LF" / "Data")
        self.current_path = self.spreadsheets_folder
        self.current_page = 0
        self.pages = []
        self.helpers = Functions()
        self.database_path = self.data_folder / "billing_accounts.sqlite3"
        self.legacy_database_paths = [
            self.data_folder / "billing_accounts",
            self.data_folder / "faturamento_contas",
        ]

    def _open_database(self):
        return SQLiteBillingDatabase(self.database_path, self.legacy_database_paths)

    def spreadsheet_interface_loop(self, system_tui):
        system_tui.current_page = 0
        while True:
            system_tui.prepare_pages(self.current_path)
            system_tui.render_tui(
                "[1] Next | [2] Previous | [F] Download Sheets | [C] Upload Sheets | "
                "[X] Create Spreadsheet | [B] Back | [ESC] Exit",
                self.current_path,
            )

            option = input("Choose an option and press Enter: ").strip().lower()
            if option == "1" and system_tui.current_page < len(system_tui.pages) - 1:
                system_tui.current_page += 1
            elif option == "2" and system_tui.current_page > 0:
                system_tui.current_page -= 1
            elif option in ["b", "s"]:
                system_tui.current_page = 0
                break
            elif option == "esc":
                sys.exit()
            elif option == "x":
                self.create_spreadsheet()
            elif option == "c":
                self.upload_spreadsheet_prompt()
            elif option == "f":
                self.download_spreadsheet_prompt()

    def upload_spreadsheet_prompt(self):
        logging.info("Starting file picker for Google Sheets upload.")
        try:
            root = Tk()
            root.withdraw()
            root.attributes("-topmost", True)

            chosen_path = filedialog.askopenfilename(
                title="Select an Excel Spreadsheet to Upload",
                filetypes=[("Excel Files", "*.xlsx *.xls")],
            )
            root.destroy()

            if chosen_path:
                file_path = Path(chosen_path)
                if file_path.exists() and file_path.suffix in [".xlsx", ".xls"]:
                    self.create_google_sheet(file_path)
                else:
                    logging.warning("The selected file is invalid.")
                    pymsgbox.alert("Invalid file or unsupported format.", "Error")
            else:
                logging.info("User canceled file selection.")
        except Exception as exc:
            logging.error("Upload processing error: %s", exc, exc_info=True)
            pymsgbox.alert(f"Upload processing error: {exc}", "Error")

    def download_spreadsheet_prompt(self):
        logging.info("Starting Google Sheets download flow.")
        archive_types = ["xlsx", "ods", "csv", "pdf", "tsv", "zip"]
        for index, available_extension in enumerate(archive_types):
            print(f" [{index}] {available_extension.upper()}")

        type_choice = input("Choose the download file type: ")
        if type_choice.isdigit() and 0 <= int(type_choice) < len(archive_types):
            selected_type = archive_types[int(type_choice)]
            url = pymsgbox.prompt("Paste the Google Sheets URL to download:")

            if url is None:
                logging.info("User canceled the URL download operation.")
            elif url.strip():
                try:
                    logging.info("URL provided for download: %s", url)
                    self.download_sheet(url, selected_type)
                except Exception as exc:
                    logging.error("Download flow error: %s", exc, exc_info=True)
                    pymsgbox.alert(f"Download error:\n{exc}", "Error")
            else:
                logging.warning("No URL was provided by the user.")
                pymsgbox.alert("Invalid or empty URL.", "Error")
        else:
            logging.warning("Invalid download file type entered in the terminal.")
            pymsgbox.alert("Invalid file type.", "Error")

    def format_worksheet(self, worksheet):
        try:
            worksheet.cell(row=1, column=1).value = "ACCOUNT"
            worksheet.cell(row=1, column=2).value = "PRODUCT ITEM"
            worksheet.cell(row=1, column=3).value = "PRICE"
            worksheet.cell(row=1, column=4).value = "REGISTRATION DATE / TIME"
            worksheet.cell(row=1, column=5).value = "CUSTOMER NAME"

            for column in range(7, 52, 4):
                cell = worksheet.cell(row=2, column=column)
                cell.value = f"--{worksheet.title.upper()} PANEL--"
            for column_id in range(9, 52, 4):
                cell = worksheet.cell(row=1, column=column_id)
                cell.value = "--POWERED BY LF SYSTEM--"
        except Exception as exc:
            logging.error("Worksheet formatting error: %s", exc)

    def create_spreadsheet(self):
        try:
            workbook = openpyxl.Workbook()
            profit_sheet = workbook.active
            profit_sheet.title = "Profit"
            workbook.create_sheet("Expenses")

            for worksheet in workbook.worksheets:
                self.format_worksheet(worksheet)

            with self._open_database() as database:
                all_accounts = database.get("accounts", database.get("contas", {}))
                if not all_accounts:
                    logging.warning("No accounts found for Excel export.")
                    pymsgbox.alert("No local database records were found to export.", "Warning")
                    return

                global_purchase_list = []
                for account_name, details in all_accounts.items():
                    history = details.get("purchase_history", details.get("historico_compras", []))
                    billing_items = details.get("billing_items", details.get("itens_faturamento", {}))
                    customer_name = billing_items.get("customer_name", billing_items.get("nome_cliente", "N/A"))

                    for purchase in history:
                        global_purchase_list.append(
                            {
                                "account": account_name,
                                "customer": customer_name,
                                "item": purchase["item"],
                                "price": purchase.get("price", purchase.get("preco")),
                                "date": purchase.get("date", purchase.get("data")),
                            }
                        )

                if not global_purchase_list:
                    logging.warning("No purchase history found for export.")
                    pymsgbox.alert("No purchase history has been registered.", "Warning")
                else:
                    ordered_purchases = sorted(
                        global_purchase_list,
                        key=lambda purchase: purchase["date"],
                        reverse=True,
                    )

                    for index, purchase in enumerate(ordered_purchases):
                        current_row = index + 3
                        profit_sheet.cell(row=current_row, column=1).value = purchase["account"]
                        profit_sheet.cell(row=current_row, column=2).value = purchase["item"]
                        profit_sheet.cell(row=current_row, column=4).value = purchase["date"]
                        profit_sheet.cell(row=current_row, column=5).value = purchase["customer"]

                        try:
                            price_cell = profit_sheet.cell(row=current_row, column=3)
                            price_cell.value = float(purchase["price"])
                            price_cell.number_format = '"$"#,##0.00'
                        except (ValueError, TypeError):
                            profit_sheet.cell(row=current_row, column=3).value = purchase["price"]

                    file_name = "LF_System_Export.xlsx"
                    save_path = self.spreadsheets_folder / file_name
                    workbook.save(save_path)
                    pymsgbox.alert(
                        f"Individual audit spreadsheet generated successfully.\nSaved as: {file_name}",
                        "Excel",
                    )
        except Exception as exc:
            logging.error("OpenPyXL error: %s", exc, exc_info=True)

    def download_sheet(self, url, archive_type):
        try:
            if "docs.google.com/spreadsheets" in url:
                url = url.split("/d/")[1].split("/edit")[0]

            sheets_workbook = ezsheets.Spreadsheet(url)
            file_name = f"Complete_Spreadsheet_{self.helpers.generate_timestamp()}.{archive_type}"
            save_path = self.spreadsheets_folder / file_name

            match archive_type.lower():
                case "xlsx":
                    sheets_workbook.downloadAsExcel(str(save_path))
                case "ods":
                    sheets_workbook.downloadAsODS(str(save_path))
                case "csv":
                    sheets_workbook.downloadAsCSV(str(save_path))
                case "tsv":
                    sheets_workbook.downloadAsTSV(str(save_path))
                case "zip":
                    sheets_workbook.downloadAsHTML(str(save_path))
                case "pdf":
                    sheets_workbook.downloadAsPDF(str(save_path))
                case _:
                    pymsgbox.alert("Invalid format.", "Error")
                    return

            pymsgbox.alert(
                f"File downloaded successfully.\nSaved as:\n{save_path.name}",
                "Success",
            )
        except Exception as exc:
            logging.error("Spreadsheet download error: %s", exc, exc_info=True)
            pymsgbox.alert(f"Spreadsheet download failed:\n\n{exc}", "Error")

    def create_google_sheet(self, path):
        try:
            logging.info("Uploading file %s to Google Sheets.", path)
            sheets_workbook = ezsheets.upload(str(path))
            logging.info("Upload completed successfully. ID: %s", sheets_workbook.id)
            pymsgbox.alert(
                text=f"Spreadsheet uploaded successfully to Google Sheets.\nTitle: {sheets_workbook.title}",
                title="Success",
            )
        except Exception as exc:
            logging.error("Google Sheets upload error: %s", exc, exc_info=True)
            pymsgbox.alert(f"Google Sheets upload failed:\n{exc}", "Error")
